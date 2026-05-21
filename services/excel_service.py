from __future__ import annotations

import shutil
from datetime import datetime, timezone
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from app.utils.config import settings
from app.utils.logger import get_logger


logger = get_logger(__name__)

EXCEL_MAX_ROWS = 1_048_576

SHEET_COLUMNS: dict[str, list[str]] = {
    "COMPANY_MASTER": [
        "Ticker",
        "NSESymbol",
        "CompanyName",
        "Exchange",
        "Sector",
        "Industry",
        "ListingDate",
        "ISIN",
        "Currency",
        "ActiveStatus",
        "FirstSeenDate",
        "LastSeenDate",
        "DataSource",
        "Notes",
    ],
    "PRICE_HISTORY": [
        "Date",
        "Ticker",
        "NSESymbol",
        "Open",
        "High",
        "Low",
        "Close",
        "AdjClose",
        "Volume",
        "Source",
        "UpdatedAt",
    ],
    "MARKET_INDEX": [
        "Date",
        "IndexTicker",
        "IndexName",
        "Open",
        "High",
        "Low",
        "Close",
        "AdjClose",
        "Volume",
        "Source",
        "UpdatedAt",
    ],
    "RISK_FREE_RATE": [
        "Date",
        "Country",
        "Tenor",
        "Rate",
        "Source",
        "UpdatedAt",
        "Notes",
    ],
    "ASSUMPTIONS": ["AssumptionName", "Value", "Source", "Notes"],
    "UPDATE_LOG": [
        "RunTimestamp",
        "Task",
        "Status",
        "RecordsAdded",
        "RecordsUpdated",
        "NewCompaniesAdded",
        "FailedTickers",
        "Notes",
    ],
    "DATA_QUALITY_CHECKS": [
        "CheckDate",
        "Ticker",
        "CheckType",
        "Status",
        "Details",
    ],
    "UNIVERSE_SNAPSHOT": [
        "SnapshotDate",
        "NSESymbol",
        "CompanyName",
        "ISIN",
        "Exchange",
        "Source",
        "Status",
    ],
}


def utc_now() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def backup_workbook() -> Path | None:
    workbook_path = settings.workbook_path
    if not workbook_path.exists():
        return None
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = settings.backup_dir / f"WACC_MARKET_DATABASE_backup_{timestamp}.xlsx"
    shutil.copy2(workbook_path, backup_path)
    logger.info("Created workbook backup at %s", backup_path)
    return backup_path


def empty_frame(sheet_name: str) -> pd.DataFrame:
    return pd.DataFrame(columns=SHEET_COLUMNS[sheet_name])


def read_sheet(sheet_name: str) -> pd.DataFrame:
    workbook_path = settings.workbook_path
    if not workbook_path.exists():
        return empty_frame(sheet_name)
    try:
        df = pd.read_excel(workbook_path, sheet_name=sheet_name, dtype=str)
    except ValueError:
        return empty_frame(sheet_name)
    return normalize_columns(sheet_name, df)


def normalize_columns(sheet_name: str, df: pd.DataFrame) -> pd.DataFrame:
    for column in SHEET_COLUMNS[sheet_name]:
        if column not in df.columns:
            df[column] = pd.NA
    return df[SHEET_COLUMNS[sheet_name]]


def merge_append_only(existing: pd.DataFrame, incoming: pd.DataFrame, keys: list[str]) -> pd.DataFrame:
    if incoming.empty:
        return existing.copy()
    merged = pd.concat([existing, incoming], ignore_index=True)
    merged = merged.drop_duplicates(subset=keys, keep="last")
    return merged.reset_index(drop=True)


def write_workbook(sheets: dict[str, pd.DataFrame]) -> Path:
    settings.generated_excel_dir.mkdir(parents=True, exist_ok=True)
    workbook_path = settings.workbook_path

    with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
        dashboard = build_dashboard(sheets)
        dashboard.to_excel(writer, sheet_name="DASHBOARD", index=False)

        for sheet_name in SHEET_COLUMNS:
            df = normalize_columns(sheet_name, sheets.get(sheet_name, empty_frame(sheet_name)))
            if sheet_name == "PRICE_HISTORY" and len(df) > EXCEL_MAX_ROWS - 1:
                write_split_price_history(writer, df)
            else:
                df.to_excel(writer, sheet_name=sheet_name, index=False)

    format_workbook(workbook_path)
    logger.info("Wrote workbook at %s", workbook_path)
    return workbook_path


def write_split_price_history(writer: pd.ExcelWriter, df: pd.DataFrame) -> None:
    chunk_size = EXCEL_MAX_ROWS - 1
    for idx, start in enumerate(range(0, len(df), chunk_size), start=1):
        df.iloc[start : start + chunk_size].to_excel(
            writer,
            sheet_name=f"PRICE_HISTORY_{idx}",
            index=False,
        )
    empty_frame("PRICE_HISTORY").to_excel(writer, sheet_name="PRICE_HISTORY", index=False)


def build_dashboard(sheets: dict[str, pd.DataFrame]) -> pd.DataFrame:
    company_master = sheets.get("COMPANY_MASTER", empty_frame("COMPANY_MASTER"))
    price_history = sheets.get("PRICE_HISTORY", empty_frame("PRICE_HISTORY"))
    update_log = sheets.get("UPDATE_LOG", empty_frame("UPDATE_LOG"))
    checks = sheets.get("DATA_QUALITY_CHECKS", empty_frame("DATA_QUALITY_CHECKS"))

    last_update = ""
    new_companies = 0
    failed_tickers = ""
    if not update_log.empty:
        latest = update_log.tail(1).iloc[0]
        last_update = latest.get("RunTimestamp", "")
        new_companies = latest.get("NewCompaniesAdded", 0)
        failed_tickers = latest.get("FailedTickers", "")

    warnings = 0
    if not checks.empty and "Status" in checks:
        warnings = int(checks["Status"].astype(str).str.upper().ne("PASS").sum())

    return pd.DataFrame(
        [
            {"Metric": "Total companies tracked", "Value": len(company_master)},
            {"Metric": "Last update timestamp", "Value": last_update},
            {"Metric": "New companies added in last run", "Value": new_companies},
            {"Metric": "Total price rows", "Value": len(price_history)},
            {"Metric": "Failed tickers in last run", "Value": failed_tickers},
            {"Metric": "Data quality warnings", "Value": warnings},
            {"Metric": "Workbook generated at UTC", "Value": utc_now()},
        ]
    )


def format_workbook(path: Path) -> None:
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        for column_cells in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in column_cells[:100])
            width = min(max(max_len + 2, 12), 45)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = width

        if ws.max_row > 1 and ws.max_column > 0:
            ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
            table_name = "".join(ch for ch in ws.title if ch.isalnum())[:25] or "Table"
            table = Table(displayName=f"{table_name}Table", ref=ref)
            style = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            table.tableStyleInfo = style
            try:
                ws.add_table(table)
            except ValueError:
                pass

    wb.save(path)


def load_all_sheets() -> dict[str, pd.DataFrame]:
    return {sheet_name: read_sheet(sheet_name) for sheet_name in SHEET_COLUMNS}

